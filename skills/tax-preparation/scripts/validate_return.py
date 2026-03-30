#!/usr/bin/env python3
"""Validate a tax computation workbook for internal consistency.

Reads the Form Values sheet from a tax computation xlsx and runs cross-form
consistency checks. Reports PASS/FAIL for each check.

Usage:
    python validate_return.py work/tax_computations.xlsx

    # Or import and use programmatically:
    from validate_return import validate_workbook
    ok, results = validate_workbook("work/tax_computations.xlsx")
"""

import argparse
import sys
import openpyxl


def read_form_values(wb):
    """Read Form Values sheet into a dict of {(form, line): value}."""
    ws = wb["Form Values"]
    values = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        form, line, desc, val = row[0], row[1], row[2], row[3]
        values[(str(form), str(line))] = val
    return values


def read_validation_sheet(wb):
    """Read the Validation sheet results."""
    if "Validation" not in wb.sheetnames:
        return []
    ws = wb["Validation"]
    results = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        results.append({
            "check": row[0],
            "expected": row[1],
            "actual": row[2],
            "status": row[3],
        })
    return results


def run_cross_checks(fv):
    """Run cross-form consistency checks on form values.

    Args:
        fv: Dict of {(form, line): value} from read_form_values()

    Returns:
        List of (description, expected, actual, passed) tuples
    """
    checks = []

    def get(form, line, default=None):
        return fv.get((form, str(line)), default)

    def check(desc, expected, actual, tol=1):
        if expected is not None and actual is not None:
            try:
                passed = abs(float(expected) - float(actual)) <= tol
            except (ValueError, TypeError):
                passed = str(expected) == str(actual)
        else:
            passed = expected is None and actual is None
        checks.append((desc, expected, actual, passed))

    # -- 1040 internal consistency --

    # Line 1z should equal line 1a (if no other 1b-1h items)
    if get("1040", "1z") and get("1040", "1a"):
        # Only check if 1z == 1a when there are no other wage adjustments
        pass  # Skip — may have tip income, dependent care, etc.

    # Line 9 = 1z + 2b + 3b + 4b + 5b + 6b + 7a + 8
    total_income_parts = sum(filter(None, [
        get("1040", "1z"), get("1040", "2b"), get("1040", "3b"),
        get("1040", "4b"), get("1040", "5b"), get("1040", "6b"),
        get("1040", "7a"), get("1040", "8"),
    ]))
    if get("1040", "9"):
        check("1040: Line 9 = sum of income lines",
              get("1040", "9"), total_income_parts)

    # Line 11 = Line 9 - Line 10
    if get("1040", "11") and get("1040", "9"):
        adj = get("1040", "10") or 0
        check("1040: Line 11 (AGI) = Line 9 - Line 10",
              get("1040", "11"), get("1040", "9") - adj)

    # Line 15 = Line 11 - Line 14
    if get("1040", "15") and get("1040", "11") and get("1040", "14"):
        check("1040: Line 15 (taxable) = Line 11 - Line 14",
              get("1040", "15"), get("1040", "11") - get("1040", "14"))

    # Line 18 = Line 16 + Line 17 (if no 17, just 16)
    if get("1040", "18") and get("1040", "16"):
        line17 = get("1040", "17") or 0
        check("1040: Line 18 = Line 16 + Line 17",
              get("1040", "18"), get("1040", "16") + line17)

    # Line 22 = Line 18 - Line 21
    if get("1040", "22") and get("1040", "18"):
        line21 = get("1040", "21") or 0
        check("1040: Line 22 = Line 18 - Line 21",
              get("1040", "22"), get("1040", "18") - line21)

    # Line 24 = Line 22 + Line 23
    if get("1040", "24") and get("1040", "22"):
        line23 = get("1040", "23") or 0
        check("1040: Line 24 (total tax) = Line 22 + Line 23",
              get("1040", "24"), get("1040", "22") + line23)

    # Line 33 = Line 25d + Line 26 + ... + Line 32
    if get("1040", "33") and get("1040", "25d"):
        payments = sum(filter(None, [
            get("1040", "25d"), get("1040", "26"),
            get("1040", "27a"), get("1040", "28"),
            get("1040", "29"), get("1040", "30"),
            get("1040", "31"), get("1040", "32"),
        ]))
        check("1040: Line 33 (total payments) = sum of payment lines",
              get("1040", "33"), payments)

    # Line 37 = Line 24 - Line 33 (amount owed)
    if get("1040", "37") and get("1040", "24") and get("1040", "33"):
        check("1040: Line 37 (owed) = Line 24 - Line 33",
              get("1040", "37"), get("1040", "24") - get("1040", "33"))

    # -- Cross-form checks --

    # Schedule D line 16 == 1040 line 7a
    if get("Schedule D", "16") and get("1040", "7a"):
        check("Schedule D line 16 == 1040 line 7a",
              get("Schedule D", "16"), get("1040", "7a"))

    # Schedule 1 line 10 == 1040 line 8
    if get("Schedule 1", "10") and get("1040", "8"):
        check("Schedule 1 line 10 == 1040 line 8",
              get("Schedule 1", "10"), get("1040", "8"))

    # Schedule 2 line 21 == 1040 line 23
    if get("Schedule 2", "21") and get("1040", "23"):
        check("Schedule 2 line 21 == 1040 line 23",
              get("Schedule 2", "21"), get("1040", "23"))

    # Schedule 3 line 8 == 1040 line 20
    if get("Schedule 3", "8") and get("1040", "20"):
        check("Schedule 3 line 8 == 1040 line 20",
              get("Schedule 3", "8"), get("1040", "20"))

    # Form 8959 line 18 == part of Schedule 2 line 11
    # Form 8960 line 17 == part of Schedule 2 line 12

    # Form 8959 line 24 == 1040 line 25c
    if get("Form 8959", "24") and get("1040", "25c"):
        check("Form 8959 line 24 (excess w/h) == 1040 line 25c",
              get("Form 8959", "24"), get("1040", "25c"))

    # Form 1116 line 35 == Schedule 3 line 1
    if get("Form 1116", "35") and get("Schedule 3", "1"):
        check("Form 1116 line 35 (FTC) == Schedule 3 line 1",
              get("Form 1116", "35"), get("Schedule 3", "1"))

    # Schedule E line 26 == Schedule 1 line 5
    if get("Schedule E", "26") and get("Schedule 1", "5"):
        check("Schedule E line 26 == Schedule 1 line 5",
              get("Schedule E", "26"), get("Schedule 1", "5"))

    # -- W-2 withholding check --
    if get("1040", "25a"):
        # 25a should match W-2 Box 2
        pass  # Requires source data — checked in the embedded validation sheet

    # -- Underpayment penalty check (Form 2210) --
    # If total tax exceeds total payments by > $1,000, the penalty computation
    # must be addressed — either Form 2210 is computed or Line 38 is explicitly
    # set (even to 0 with a note). Skipping this lets the IRS compute it at a
    # higher amount.
    total_tax = get("1040", "24")
    total_payments = get("1040", "33")
    penalty_line = get("1040", "38")
    if total_tax is not None and total_payments is not None:
        balance_due = total_tax - total_payments
        if balance_due > 1000:
            check("1040: Underpayment penalty (Line 38 / Form 2210) computed when balance due > $1,000",
                  True, penalty_line is not None)

    return checks


def run_input_checks(wb, forms_dir=None):
    """Run input-level validation: verify Tax Tables values against forms.

    Unlike cross-form checks (which verify internal consistency), these
    checks verify that the INPUT values are correct by comparing them
    against values extracted directly from the downloaded PDF forms.

    Args:
        wb: openpyxl workbook
        forms_dir: Path to forms/ directory. If None, skips form extraction checks.

    Returns:
        List of (description, expected, actual, passed) tuples
    """
    checks = []

    def check(desc, expected, actual, tol=1):
        if expected is not None and actual is not None:
            try:
                passed = abs(float(expected) - float(actual)) <= tol
            except (ValueError, TypeError):
                passed = str(expected) == str(actual)
        else:
            passed = expected is None and actual is None
        checks.append((desc, expected, actual, passed))

    # Read Tax Tables sheet
    tax_tables = {}
    if "Tax Tables" in wb.sheetnames:
        ws = wb["Tax Tables"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1] and row[2] is not None:
                tax_tables[(str(row[0]), str(row[1]))] = row[2]

    # If forms_dir provided, cross-check against extracted values
    if forms_dir:
        try:
            from extract_tax_tables import extract_all
            extracted = extract_all(forms_dir)

            fed = extracted.get("federal", {})
            ca = extracted.get("ca_540", {})
            ca_ca = extracted.get("ca_schedule_ca", {})

            # Federal standard deduction
            if "standard_deduction_single" in fed:
                # Find the standard deduction in the workbook's Form Values
                fv = read_form_values(wb) if "Form Values" in wb.sheetnames else {}
                fed_ded = fv.get(("1040", "12"), fv.get(("1040", "12e")))
                if fed_ded:
                    check("INPUT: Federal deduction ≤ form's std ded (or is itemized)",
                          True, fed_ded <= fed["standard_deduction_single"] or fed_ded > fed["standard_deduction_single"])
                    # The real check: if using standard, it must match the form
                    # If itemized, it should be > standard (otherwise why itemize?)

                check("INPUT: Federal std ded on form matches Tax Tables",
                      fed["standard_deduction_single"],
                      tax_tables.get(("F", "sd"), tax_tables.get(("Federal Brackets (Single)", "Standard deduction"))))

            # CA standard deduction
            if "standard_deduction_single" in ca:
                check("INPUT: CA std ded on form matches Tax Tables",
                      ca["standard_deduction_single"],
                      tax_tables.get(("C", "sd"), tax_tables.get(("CA (Single)", "Standard deduction"))))

            # CA Pease threshold
            if "pease_threshold_single" in ca_ca:
                check("INPUT: CA Pease threshold on form matches Tax Tables",
                      ca_ca["pease_threshold_single"],
                      tax_tables.get(("C", "pease"), tax_tables.get(("CA (Single)", "Pease threshold")))
                )

            # CA exemption credit
            if "dependent_exemption_credit" in ca:
                check("INPUT: CA exemption credit on form matches Tax Tables",
                      ca["dependent_exemption_credit"],
                      tax_tables.get(("C", "exemption"), tax_tables.get(("CA (Single)", "Exemption credit")))
                )

        except ImportError:
            checks.append(("INPUT: extract_tax_tables.py available", True, False, False))
        except Exception as e:
            checks.append((f"INPUT: form extraction error: {e}", True, False, False))

    # Currency consistency: if FTC exists, check that foreign amounts are reasonable
    fv = read_form_values(wb) if "Form Values" in wb.sheetnames else {}
    ftc = fv.get(("Form 1116", "35"), fv.get(("Form 1116", "24")))
    if ftc and ftc > 0:
        # FTC should generally be less than total tax
        total_tax = fv.get(("1040", "24"), fv.get(("1040", "16")))
        if total_tax:
            check("INPUT: FTC ≤ total tax (sanity)",
                  True, ftc <= total_tax)

    # CA AGI vs Federal AGI: if HSA or other state adjustments exist,
    # CA AGI should differ from federal AGI
    fed_agi = fv.get(("1040", "11"))
    ca_agi = fv.get(("CA 540", "17"), fv.get(("CA540", "17")))
    # Note: we can't automatically know if HSA applies without source data,
    # but we can flag if they're identical when the user reported HSA
    # This check is informational — the embedded validation should catch specifics

    # Source document reconciliation: verify Source Data totals match computations
    source_checks = run_source_reconciliation(wb)
    checks.extend(source_checks)

    return checks


def run_source_reconciliation(wb):
    """Reconcile Source Data totals against computation sheet totals.

    Checks that every dollar amount extracted from source documents is
    accounted for in the computation sheets. Catches data that was
    extracted but dropped, or never extracted at all.

    Looks for document-level totals in Source Data (rows with 'total'
    in the item name) and compares against the sum of individual items
    from that same source.
    """
    checks = []

    if "Source Data" not in wb.sheetnames:
        return checks

    ws = wb["Source Data"]

    # Collect all source data rows: {category: [(item, value, source), ...]}
    by_category = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        category, item, value, source = (
            str(row[0]) if row[0] else "",
            str(row[1]) if row[1] else "",
            row[2],
            str(row[3]) if len(row) > 3 and row[3] else "",
        )
        if category not in by_category:
            by_category[category] = []
        by_category[category].append((item, value, source))

    # For each category, check if there's a "total" row and if individual
    # items sum to it
    for category, rows in by_category.items():
        total_rows = [(item, val) for item, val, _ in rows
                      if 'total' in item.lower() and val is not None
                      and isinstance(val, (int, float))]
        if not total_rows:
            continue

        for total_item, total_val in total_rows:
            # Find the keyword that identifies what this total covers
            # e.g., "Total Proceeds" -> look for other rows with "Proceeds"
            total_lower = total_item.lower()
            keyword = None
            for kw in ['proceeds', 'cost', 'basis', 'gain', 'loss',
                       'income', 'dividend', 'interest', 'rent',
                       'tax', 'withheld', 'withholding']:
                if kw in total_lower:
                    keyword = kw
                    break

            if keyword is None:
                continue

            # Sum individual (non-total) rows that contain the same keyword
            parts = []
            for item, val, _ in rows:
                if 'total' not in item.lower() and keyword in item.lower() \
                        and val is not None and isinstance(val, (int, float)):
                    parts.append(val)

            if parts:
                parts_sum = sum(parts)
                checks.append((
                    f"SOURCE: {category} — {total_item} ({total_val:,.2f}) "
                    f"== sum of {len(parts)} items ({parts_sum:,.2f})",
                    total_val, parts_sum,
                    abs(total_val - parts_sum) <= 1
                ))

    return checks


def validate_workbook(path, forms_dir=None):
    """Validate a tax computation workbook.

    Args:
        path: Path to tax_computations.xlsx
        forms_dir: Optional path to forms/ directory for input-level checks

    Returns:
        (all_passed: bool, results: list of (desc, expected, actual, passed))
    """
    wb = openpyxl.load_workbook(path, data_only=True)

    results = []
    all_passed = True

    # Check embedded validation sheet
    embedded = read_validation_sheet(wb)
    for item in embedded:
        passed = item["status"] == "PASS"
        results.append((item["check"], item["expected"], item["actual"], passed))
        if not passed:
            all_passed = False

    # Run input-level checks (compare Tax Tables against forms)
    input_checks = run_input_checks(wb, forms_dir)
    for desc, expected, actual, passed in input_checks:
        results.append((desc, expected, actual, passed))
        if not passed:
            all_passed = False

    # Run cross-form checks
    if "Form Values" in wb.sheetnames:
        fv = read_form_values(wb)
        cross_checks = run_cross_checks(fv)
        for desc, expected, actual, passed in cross_checks:
            results.append((desc, expected, actual, passed))
            if not passed:
                all_passed = False

    return all_passed, results


def main():
    parser = argparse.ArgumentParser(description="Validate tax computation workbook")
    parser.add_argument("workbook", help="Path to tax_computations.xlsx")
    parser.add_argument("--forms-dir", help="Path to forms/ directory for input-level checks")
    args = parser.parse_args()

    all_passed, results = validate_workbook(args.workbook, args.forms_dir)

    print(f"\nValidation Results for: {args.workbook}")
    print("=" * 70)

    for desc, expected, actual, passed in results:
        status = "PASS" if passed else "FAIL"
        symbol = "✓" if passed else "✗"
        print(f"  {symbol} {status}: {desc}")
        if not passed:
            print(f"         Expected: {expected}")
            print(f"         Actual:   {actual}")

    print("=" * 70)
    total = len(results)
    passed = sum(1 for _, _, _, p in results if p)
    failed = total - passed
    print(f"  {passed}/{total} checks passed", end="")
    if failed > 0:
        print(f", {failed} FAILED")
    else:
        print()

    # Reminder: re-read SKILL.md to catch any missed steps.
    # This prints unconditionally because validation is the gate between
    # Phase 2 (computation) and Phase 3 (form filling). The agent's context
    # may have been compacted by this point, losing earlier instructions.
    print()
    print("=" * 70)
    print("CHECKPOINT: Re-read SKILL.md for the tax-preparation skill NOW.")
    print("Verify you have not skipped any steps, including but not limited to:")
    print("  - Underpayment penalty (Form 2210) — Phase 2")
    print("  - Other obligations (FBAR, FATCA, Form 2210 safe harbor) — Phase 3d")
    print("  - Verify against form instructions (fetch + read) — Phase 3c")
    print("  - State return non-conformity items — Phase 2 state section")
    print("  - Carryforwards sheet for next year — Phase 2 workbook")
    print("=" * 70)

    sys.exit(1 if failed > 0 else 0)


if __name__ == "__main__":
    main()
